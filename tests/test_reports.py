from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from engine.checks import run_checks
from engine.constants import EXPORT_DISCLAIMER
from engine.report_excel import build_excel_working_paper
from engine.report_pdf import build_pdf_report


def test_excel_export_creates_required_sheets(demo_inputs):
    findings = run_checks(demo_inputs, today=date(2026, 7, 4))
    workbook = load_workbook(BytesIO(build_excel_working_paper(demo_inputs, findings)), read_only=True)
    assert workbook.sheetnames == [
        "Summary",
        "VAT Returns",
        "P&L Summary",
        "CT Computation",
        "Reconciling Items",
        "Checks",
        "Methodology",
    ]


def test_pdf_export_returns_bytes_and_includes_disclaimer(demo_inputs):
    findings = run_checks(demo_inputs, today=date(2026, 7, 4))
    pdf = build_pdf_report(demo_inputs, findings)
    assert pdf.startswith(b"%PDF")
    assert EXPORT_DISCLAIMER.split(".")[0].encode() in pdf

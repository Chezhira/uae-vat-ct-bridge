from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook

from engine.models import PLSummary, ReconcilingItem, VATReturnPeriod

VAT_COLUMNS = [
    "period_start",
    "period_end",
    "std_rated_supplies",
    "zero_rated_supplies",
    "exempt_supplies",
    "reverse_charge_supplies",
    "goods_imported",
    "adjustments",
    "recoverable_input_vat_base",
    "output_vat",
    "input_vat",
]
PL_COLUMNS = ["accounting_revenue", "other_income", "gains_on_disposal", "total_expenses", "accounting_profit"]
RECON_COLUMNS = ["category", "description", "amount", "evidence_ref"]


def _read_table(source: str | Path | BinaryIO | BytesIO, filename: str | None = None) -> pd.DataFrame:
    name = filename or str(source)
    suffix = Path(name).suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(source)
    if suffix == ".xlsx":
        return _read_xlsx_values(source)
    if suffix == ".xlsm":
        raise ValueError("XLSM files are not accepted.")
    raise ValueError("Only CSV and XLSX files are accepted.")


def _read_xlsx_values(source: str | Path | BinaryIO | BytesIO) -> pd.DataFrame:
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return pd.DataFrame()
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    data_rows = [list(row) for row in rows[1:] if any(value is not None for value in row)]
    return pd.DataFrame(data_rows, columns=headers)


def _require_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    missing = [column for column in columns if column not in df.columns]
    if missing:
        raise ValueError("Missing required columns: " + ", ".join(missing))
    return df[columns]


def parse_vat_returns(source: str | Path | BinaryIO | BytesIO, filename: str | None = None) -> list[VATReturnPeriod]:
    df = _require_columns(_read_table(source, filename), VAT_COLUMNS)
    return [VATReturnPeriod.model_validate(record) for record in df.to_dict(orient="records")]


def parse_pl_summary(source: str | Path | BinaryIO | BytesIO, filename: str | None = None) -> PLSummary:
    df = _require_columns(_read_table(source, filename), PL_COLUMNS)
    if len(df) != 1:
        raise ValueError("P&L summary must contain exactly one row.")
    return PLSummary.model_validate(df.iloc[0].to_dict())


def parse_reconciling_items(
    source: str | Path | BinaryIO | BytesIO,
    filename: str | None = None,
) -> list[ReconcilingItem]:
    df = _require_columns(_read_table(source, filename), RECON_COLUMNS).fillna("")
    return [ReconcilingItem.model_validate(record) for record in df.to_dict(orient="records")]
